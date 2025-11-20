import os
from openpyxl import load_workbook
import sys
import re
import subprocess
import tempfile
import shutil

print(sys.executable)

# 🔧 SETTINGS – edit these before running
root_folder = r"C:\Users\judep\Downloads\FORMS EDITING\1. Accounting Forms"  # Folder containing Excel files
find_text = "Belships"
replace_text = "GMSMI"
match_variations = True  # If True, finds case-insensitive and whitespace variations

# --- Do not edit below this line ---
count_files = 0
count_replaced = 0
count_converted = 0

def convert_xls_to_xlsx(xls_path):
    """Convert .xls or .xlsm to .xlsx using LibreOffice or Excel COM."""
    try:
        # Try LibreOffice first
        xlsx_path = xls_path.rsplit(".", 1)[0] + ".xlsx"
        subprocess.run([
            "soffice", "--headless", "--convert-to", "xlsx", "--outdir",
            os.path.dirname(xls_path), xls_path
        ], capture_output=True, timeout=30)
        
        if os.path.exists(xlsx_path):
            return xlsx_path
    except Exception:
        pass
    
    # Try Excel COM (Windows only)
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        xlsx_path = xls_path.rsplit(".", 1)[0] + ".xlsx"
        wb = excel.Workbooks.Open(os.path.abspath(xls_path))
        wb.SaveAs(os.path.abspath(xlsx_path), FileFormat=51)  # 51 = .xlsx
        wb.Close()
        excel.Quit()
        return xlsx_path
    except Exception:
        pass
    
    return None

def create_variation_pattern(text):
    """Create a regex pattern that matches variations of the text."""
    # Escape special regex characters
    escaped = re.escape(text)
    # Allow optional whitespace variations (spaces, tabs, etc.)
    pattern = re.sub(r'\\s+', r'\\s+', escaped)
    # Make it case-insensitive
    return re.compile(pattern, re.IGNORECASE)

for foldername, subfolders, filenames in os.walk(root_folder):
    for filename in filenames:
        # Skip temporary Excel files
        if filename.startswith("~$"):
            continue
        
        # Process .xlsx, .xls, and .xlsm files
        if filename.endswith((".xlsx", ".xls", ".xlsm")):
            file_path = os.path.join(foldername, filename)
            
            # Convert old formats to xlsx
            if filename.endswith((".xls", ".xlsm")) and not filename.endswith(".xlsx"):
                print(f"🔄 Converting {filename} to .xlsx...")
                converted_path = convert_xls_to_xlsx(file_path)
                if converted_path and os.path.exists(converted_path):
                    file_path = converted_path
                    count_converted += 1
                    print(f"✅ Converted: {converted_path}")
                else:
                    print(f"⚠️ Could not convert {filename}. Skipping.")
                    continue
            
            try:
                wb = load_workbook(file_path)
            except Exception as e:
                print(f"⚠️ Skipped {filename} (error reading file: {e})")
                continue

            replaced_in_file = False
            
            # Create pattern for text variations if enabled
            if match_variations:
                pattern = create_variation_pattern(find_text)
            else:
                pattern = None

            # Loop through all sheets and cells
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            if pattern:
                                # Use regex to match variations
                                if re.search(pattern, cell.value):
                                    cell.value = re.sub(pattern, replace_text, cell.value)
                                    replaced_in_file = True
                            else:
                                # Exact match
                                if find_text in cell.value:
                                    cell.value = cell.value.replace(find_text, replace_text)
                                    replaced_in_file = True

            if replaced_in_file:
                wb.save(file_path)
                count_replaced += 1
                print(f"✅ Modified: {file_path}")
            else:
                print(f"— No change: {file_path}")

            count_files += 1

print(f"\n✅ Processed {count_files} Excel files (including subfolders).")
print(f"🔄 Converted {count_converted} old Excel files to .xlsx.")
print(f"📝 Updated {count_replaced} files containing '{find_text}'.")
if match_variations:
    print(f"   (Found text variations: case-insensitive + whitespace variations)")
print("🎉 Done!")
